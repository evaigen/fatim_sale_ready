import requests
from lxml import etree
from openpyxl import load_workbook, Workbook
from openpyxl.styles.colors import WHITE, RGB
import warnings

# !/usr/bin/env python3
# -*- coding: utf-8 -*-

# FIXING THE ARGB HEX VALUES ERROR
__old_rgb_set__ = RGB.__set__

# Dictionary of markings
fatim_codes = {
               'arm': ['Зикрач Армавир', 0.03, 4.0, 1.02],
               'shev': ['Шевченко Shev', 0.03, 4.0, 1.02],
               'volg': ['Копач', 0.03, 4.0, 1.02],
               'misha': ['Шевченко Misha', 0.03, 4.0, 1.02],
               'kisa': ['Зикрач Краснодар', 0.03, 4.0, 1.02]
}

# XPath expression to extract the currency rate
x_dollar = '//*[@id="content"]/div/div/div/div[3]/div/table/tbody/tr[15]/td[5]'
# x_euro = '//*[@id="content"]/div/div/div/div[3]/div/table/tbody/tr[16]/td[5]'


def __rgb_set_fixed__(self, instance, value):
    try:
        __old_rgb_set__(self, instance, value)
    except ValueError as e:
        if e.args[0] == 'Colors must be aRGB hex values':
            __old_rgb_set__(self, instance, WHITE)


def parsing_currency():
    # URL of the currency website
    url = "https://www.cbr.ru/currency_base/daily/"

    # Send a GET request to the website
    response = requests.get(url)

    # Check if the request was successful
    if response.status_code == 200:
        # Parse the HTML content
        html_content = response.text
        parser = etree.HTMLParser()
        tree = etree.fromstring(html_content, parser)

        # Find the currency rate element using XPath
        dollar_element = tree.xpath(x_dollar)
        # euro_element = tree.xpath(x_euro)

        # Check if the element was found
        if dollar_element:
            # Get the text content of the element
            dollar_rate = dollar_element[0].text
            # euro_rate = euro_element[0].text

            dollar_rate = dollar_rate.replace(",", ".")
            # euro_rate = euro_rate.replace(",", ".")
            # dollar_rate = (float(dollar_rate) + markup_rub)*markup_percent
            # euro_rate = (float(euro_rate) + markup_rub)*markup_percent

            # Print the currency rate
            print(f"USD to RUB currency rate: {dollar_rate}")
            # print(f"USD to RUB currency rate: {euro_rate}")
            return float(dollar_rate)
        else:
            print("Currency rate element not found.")
    else:
        print("Failed to fetch the webpage.")
        exit()


def load_invoice(fatim_path):
    try:
        fatim_workbook = load_workbook(fatim_path)
        fatim_worksheet = fatim_workbook.active
        total_rows = fatim_worksheet.max_row

        for row in reversed(list(fatim_worksheet.iter_rows(min_row=1,
                                                           max_row=total_rows,
                                                           min_col=1,
                                                           max_col=1))):
            if row[0].value is None:
                fatim_worksheet.delete_rows(row[0].row, 1)

    except Exception as e:
        print(f"Error has occured: {e}")
        exit()

    return fatim_workbook, fatim_worksheet


def fatim_upd(fatim_worksheet, dollar_rate):
    total_rows = fatim_worksheet.max_row + 1
    flower_sum = {
                  'shev': [0.0, 0.0, 0.0, 0.0], 'arm': [0.0, 0.0, 0.0, 0.0],
                  'volg': [0.0, 0.0, 0.0, 0.0], 'kisa': [0.0, 0.0, 0.0, 0.0],
                  'misha': [0.0, 0.0, 0.0, 0.0]
    }

    for row in range(2, total_rows):
        code_name = str(fatim_worksheet[f'B{row}'].value).lower()
        flower_type = str(fatim_worksheet[f'D{row}'].value).lower()
        flower_price = float(fatim_worksheet[f'H{row}'].value)
        flower_amount = float(fatim_worksheet[f'G{row}'].value)
        markup = float(fatim_worksheet[f'I{row}'].value)

        for code, info in fatim_codes.items():
            if code in code_name:
                dollar_rate_upd = (float(dollar_rate) + info[2])*info[3]
                fatim_worksheet[f'L{row}'] = dollar_rate_upd

                if 'rose' in flower_type:
                    fatim_worksheet[f'I{row}'] = markup + info[1]

                currency_rub = float(fatim_worksheet[f'L{row}'].value)
                markup = float(fatim_worksheet[f'I{row}'].value)

                new_sum = ((flower_price + markup) * flower_amount)
                fatim_worksheet[f'K{row}'] = new_sum
                flower_usd = float(fatim_worksheet[f'K{row}'].value)
                new_sum = new_sum * currency_rub
                fatim_worksheet[f'M{row}'] = new_sum
                flower_rub = float(fatim_worksheet[f'M{row}'].value)
                flower_sum[code][0] = flower_sum[code][0] + flower_rub
                flower_sum[code][3] = flower_sum[code][3] + flower_usd
                fatim_worksheet[f'J{row}'] = flower_price + markup

                break

    for code in flower_sum.keys():
        if flower_sum[code][0] != 0.0:
            truck_cost = float(input(f'Total logistics for code {code}:\n'))
            flower_sum[code][1] = truck_cost
            flower_sum[code][2] = round(flower_sum[code][1]/flower_sum[code][0], 8)

    for row in range(2, total_rows):
        flower_amount = float(fatim_worksheet[f'G{row}'].value)
        flower_rub = float(fatim_worksheet[f'M{row}'].value)
        code_name = str(fatim_worksheet[f'B{row}'].value).lower()

        for code in fatim_codes.keys():
            if code in code_name:
                fatim_worksheet[f'N{row}'] = flower_sum[code][2] * flower_rub
                truck_rub = float(fatim_worksheet[f'N{row}'].value)
                fatim_worksheet[f'O{row}'] = truck_rub + flower_rub
                total_rub = float(fatim_worksheet[f'O{row}'].value)
                fatim_worksheet[f'P{row}'] = total_rub/flower_amount

    total_sale = 0.0

    for info in flower_sum.values():
        total_sale += info[3]

    print(f'Total flower sale in USD: {total_sale}')


def fatim_divide(fatim_worksheet, code, info):
    total_rows = fatim_worksheet.max_row + 1
    new_wb = Workbook()
    new_sheet = new_wb.active
    start = 0
    finish = 0

    for row in range(2, total_rows):
        code_name = str(fatim_worksheet[f'B{row}'].value).lower()

        if (
            code in code_name
            and
            start == 0
        ):
            start = row

        elif (
            code not in code_name
            and
            start != 0
        ):
            finish = row - 1
            break

        elif (
            row == total_rows - 1
            and
            start != 0
        ):
            finish = row

    if start and finish:
        new_row = 1
        for row in range(start, finish + 1):
            for col in range(1, 17):
                cell_value = fatim_worksheet.cell(row=row, column=col).value
                new_sheet.cell(row=new_row, column=col, value=cell_value)
            new_row += 1

        new_wb.save(f'Fatim {info[0]}.xlsx')


def start():
    invoice_name = input('Name of the invoice:\n')
    fatim_path = f'{invoice_name}.xlsx'
    fatim_workbook, fatim_worksheet = load_invoice(fatim_path)
    dollar_rate = parsing_currency()
    fatim_upd(fatim_worksheet, dollar_rate)

    for code, info in fatim_codes.items():
        fatim_divide(fatim_worksheet, code, info)


warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
RGB.__set__ = __rgb_set_fixed__
start()
